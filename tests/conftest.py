"""
共有フィクスチャ。
openpyxl でインメモリの .xlsx ファイルを生成し BytesIO として返す。
"""

import datetime
from io import BytesIO

import openpyxl
import pytest


def _to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# simple_wb
#   Sheet1: A列ラベル・B列値の典型的フォーム
#   A1:氏名  B1:山田太郎
#   A2:年齢  B2:30
#   A3:メール B3:yamada@example.com
#   A4:部署  B4:営業部
# ---------------------------------------------------------------------------
@pytest.fixture
def simple_wb() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "氏名"
    ws["B1"] = "山田太郎"
    ws["A2"] = "年齢"
    ws["B2"] = 30
    ws["A3"] = "メール"
    ws["B3"] = "yamada@example.com"
    ws["A4"] = "部署"
    ws["B4"] = "営業部"
    return _to_bytes(wb)


# ---------------------------------------------------------------------------
# nested_wb
#   Sheet1: 住所などネスト構造向けデータ
#   B1:山田太郎  B2:東京都  B3:渋谷区  B4:150-0002  B5:03-1234-5678
# ---------------------------------------------------------------------------
@pytest.fixture
def nested_wb() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B1"] = "山田太郎"
    ws["B2"] = "東京都"
    ws["B3"] = "渋谷区"
    ws["B4"] = "150-0002"
    ws["B5"] = "03-1234-5678"
    return _to_bytes(wb)


# ---------------------------------------------------------------------------
# range_wb
#   Sheet1: A1:A5 に果物リスト
#   A1:りんご A2:みかん A3:ぶどう A4:もも A5:なし
# ---------------------------------------------------------------------------
@pytest.fixture
def range_wb() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    fruits = ["りんご", "みかん", "ぶどう", "もも", "なし"]
    for i, name in enumerate(fruits, start=1):
        ws[f"A{i}"] = name
    return _to_bytes(wb)


# ---------------------------------------------------------------------------
# table_wb
#   Sheet1: A1:C1 ヘッダー、A2:C4 データ（商品テーブル）
#   A1:id  B1:name  C1:price
#   A2:1   B2:りんご C2:100
#   A3:2   B3:みかん C3:80
#   A4:3   B4:ぶどう C4:200
# ---------------------------------------------------------------------------
@pytest.fixture
def table_wb() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["id", "name", "price"]
    rows = [(1, "りんご", 100), (2, "みかん", 80), (3, "ぶどう", 200)]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, (id_, name, price) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=id_)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=price)
    return _to_bytes(wb)


# ---------------------------------------------------------------------------
# table_with_empty_wb
#   Sheet1: A1:C6 データ。3行目が空行（skip_empty テスト用）
#   A2:1 B2:りんご C2:100
#   A3:  B3:       C3:       ← 空行
#   A4:2 B4:みかん C4:80
# ---------------------------------------------------------------------------
@pytest.fixture
def table_with_empty_wb() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A2"] = 1
    ws["B2"] = "りんご"
    ws["C2"] = 100
    # row 3 is empty
    ws["A4"] = 2
    ws["B4"] = "みかん"
    ws["C4"] = 80
    return _to_bytes(wb)


# ---------------------------------------------------------------------------
# column_range_wb
#   Sheet1: A1:D1 に横方向のデータ（direction=column テスト用）
#   A1:春 B1:夏 C1:秋 D1:冬
# ---------------------------------------------------------------------------
@pytest.fixture
def column_range_wb() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col, season in enumerate(["春", "夏", "秋", "冬"], start=1):
        ws.cell(row=1, column=col, value=season)
    return _to_bytes(wb)


# ---------------------------------------------------------------------------
# multi_sheet_wb
#   「顧客情報」シート: B1:C-001 B2:山田太郎 B3:yamada@example.com
#   「注文情報」シート: B1:ORD-001 B2:2024-01-15（日付） B5:15000
# ---------------------------------------------------------------------------
@pytest.fixture
def multi_sheet_wb() -> bytes:
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "顧客情報"
    ws1["B1"] = "C-001"
    ws1["B2"] = "山田太郎"
    ws1["B3"] = "yamada@example.com"

    ws2 = wb.create_sheet("注文情報")
    ws2["B1"] = "ORD-001"
    ws2["B2"] = datetime.datetime(2024, 1, 15)
    ws2["B5"] = 15000

    return _to_bytes(wb)


# ---------------------------------------------------------------------------
# types_wb
#   Sheet1: 各種セル型のテスト
#   A1:42（整数）  A2:3.14（小数）  A3:hello（文字列）
#   A4:True（真偽値）  A5:datetime（日付）  A6:None（空白）
# ---------------------------------------------------------------------------
@pytest.fixture
def types_wb() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 42
    ws["A2"] = 3.14
    ws["A3"] = "hello"
    ws["A4"] = True
    ws["A5"] = datetime.datetime(2024, 6, 1, 12, 0, 0)
    # A6 は意図的に空白（何も書かない）
    return _to_bytes(wb)


# ---------------------------------------------------------------------------
# empty_cell_wb
#   Sheet1: B1 のみ値あり。B2・B3 は空白（empty_cell オプションのテスト用）
# ---------------------------------------------------------------------------
@pytest.fixture
def empty_cell_wb() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B1"] = "存在する値"
    # B2, B3 は空白
    return _to_bytes(wb)


# ---------------------------------------------------------------------------
# second_sheet_wb
#   Sheet1（インデックス0）: A1:シート1の値
#   Sheet2（インデックス1）: A1:シート2の値
# ---------------------------------------------------------------------------
@pytest.fixture
def second_sheet_wb() -> bytes:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "シート1の値"
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "シート2の値"
    return _to_bytes(wb)
