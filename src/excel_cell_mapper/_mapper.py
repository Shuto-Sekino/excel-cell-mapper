from __future__ import annotations

import datetime
import io
from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO, Callable, Literal

import openpyxl
from openpyxl.utils import get_column_letter

from excel_cell_mapper._cell_ref import (
    RangeAddress,
    is_bare_cell_ref,
    parse_cell_ref,
    parse_range_ref,
)
from excel_cell_mapper._exceptions import (
    InvalidSchemaError,
    ParseError,
    SheetNotFoundError,
)

CellValue = str | int | float | bool | datetime.datetime | None
ExcelSource = str | Path | bytes | BinaryIO


@dataclass
class CellContext:
    cell_ref: str
    sheet_name: str
    col_index: int  # 0-based
    row_index: int  # 0-based


CellTransformer = Callable[[CellValue, CellContext], object]

_EMPTY_CELL_OPTIONS = {"none", "omit", "empty"}
_DATE_FORMAT_OPTIONS = {"datetime", "iso", "local"}


class ExcelMapper:
    def __init__(
        self,
        source: ExcelSource,
        *,
        default_sheet: str | int | None = None,
        empty_cell: Literal["none", "omit", "empty"] = "none",
        date_format: Literal["datetime", "iso", "local"] = "datetime",
        transform: CellTransformer | None = None,
    ) -> None:
        if empty_cell not in _EMPTY_CELL_OPTIONS:
            raise ValueError(
                f"empty_cell must be one of {_EMPTY_CELL_OPTIONS}, got {empty_cell!r}"
            )
        if date_format not in _DATE_FORMAT_OPTIONS:
            raise ValueError(
                f"date_format must be one of {_DATE_FORMAT_OPTIONS}, got {date_format!r}"
            )

        self._wb = self._load_workbook(source)
        self._empty_cell = empty_cell
        self._date_format = date_format
        self._transform = transform

        # Validate and resolve default_sheet early so errors surface at init time.
        if default_sheet is not None:
            self._resolve_sheet(default_sheet)
        self._default_sheet = default_sheet

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def map(self, schema: object, *, sheet: str | int | None = None) -> dict:
        effective_sheet = sheet if sheet is not None else self._default_sheet
        if effective_sheet is not None:
            self._resolve_sheet(effective_sheet)  # validate early
        result = self._resolve_schema(schema, effective_sheet)
        if not isinstance(result, dict):
            raise InvalidSchemaError("Top-level schema must be a dict.")
        return result

    def get_cell(self, cell_ref: str) -> CellValue:
        addr = parse_cell_ref(cell_ref)
        ws = self._get_worksheet(addr.sheet or self._default_sheet_name())
        return self._cell_value(ws, addr.row, addr.col, cell_ref)

    def get_range(self, range_ref: str) -> list[list[CellValue]]:
        addr = parse_range_ref(range_ref)
        ws = self._get_worksheet(addr.sheet or self._default_sheet_name())
        return self._read_range_2d(ws, addr)

    def get_sheet_names(self) -> list[str]:
        return self._wb.sheetnames

    # ------------------------------------------------------------------
    # Context manager
    # ------------------------------------------------------------------

    def __enter__(self) -> ExcelMapper:
        return self

    def __exit__(self, *_) -> None:
        self._wb.close()

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _load_workbook(source: ExcelSource) -> openpyxl.Workbook:
        try:
            if isinstance(source, (str, Path)):
                return openpyxl.load_workbook(source, data_only=True)
            if isinstance(source, bytes):
                return openpyxl.load_workbook(io.BytesIO(source), data_only=True)
            # BinaryIO (file-like)
            return openpyxl.load_workbook(source, data_only=True)
        except Exception as exc:
            raise ParseError(f"Failed to parse Excel source: {exc}") from exc

    def _resolve_sheet(self, sheet: str | int) -> str:
        """Return sheet name for *sheet* (name or 0-based index)."""
        names = self._wb.sheetnames
        if isinstance(sheet, int):
            if sheet < 0 or sheet >= len(names):
                raise SheetNotFoundError(str(sheet))
            return names[sheet]
        if sheet not in names:
            raise SheetNotFoundError(sheet)
        return sheet

    def _default_sheet_name(self) -> str:
        if self._default_sheet is None:
            return self._wb.sheetnames[0]
        return self._resolve_sheet(self._default_sheet)

    def _get_worksheet(self, sheet_name: str):
        if sheet_name not in self._wb.sheetnames:
            raise SheetNotFoundError(sheet_name)
        return self._wb[sheet_name]

    def _cell_value(self, ws, row: int, col: int, cell_ref: str) -> CellValue:
        raw = ws.cell(row=row, column=col).value
        value = self._convert_value(raw)
        if self._transform is not None:
            col_letter = get_column_letter(col)
            ref_bare = f"{col_letter}{row}"
            ctx = CellContext(
                cell_ref=ref_bare,
                sheet_name=ws.title,
                col_index=col - 1,
                row_index=row - 1,
            )
            return self._transform(value, ctx)  # type: ignore[return-value]
        return value

    def _convert_value(self, raw) -> CellValue:
        if raw is None:
            return None
        if isinstance(raw, bool):
            return raw
        if isinstance(raw, int):
            return raw
        if isinstance(raw, float):
            # openpyxl sometimes returns integer-valued floats for int cells
            if raw == int(raw):
                return int(raw)
            return raw
        if isinstance(raw, datetime.datetime):
            return self._convert_date(raw)
        if isinstance(raw, datetime.date):
            return self._convert_date(datetime.datetime(raw.year, raw.month, raw.day))
        return str(raw)

    def _convert_date(self, dt: datetime.datetime) -> CellValue:
        if self._date_format == "datetime":
            return dt
        if self._date_format == "iso":
            return dt.isoformat()
        # "local"
        return str(dt)

    def _read_range_2d(self, ws, addr: RangeAddress) -> list[list[CellValue]]:
        result = []
        for row in range(addr.row_start, addr.row_end + 1):
            row_data = []
            for col in range(addr.col_start, addr.col_end + 1):
                row_data.append(
                    self._cell_value(ws, row, col, f"{get_column_letter(col)}{row}")
                )
            result.append(row_data)
        return result

    # ------------------------------------------------------------------
    # Schema resolution
    # ------------------------------------------------------------------

    _DIRECTIVES = {"$range", "$schema", "$direction", "$skip_empty"}

    @staticmethod
    def _is_cell_object(schema: dict) -> bool:
        """{"cell": "B1"} や {"cell": "B1", "sheet": "..."} 形式かどうか判定する。"""
        return "cell" in schema and schema.keys() <= {"cell", "sheet"}

    def _resolve_schema(
        self, schema: object, default_sheet: str | int | None
    ) -> object:
        if isinstance(schema, str):
            return self._resolve_cell_ref_value(schema, default_sheet)

        if isinstance(schema, list):
            return self._resolve_list_schema(schema, default_sheet)

        if isinstance(schema, dict):
            if "$range" in schema or "$schema" in schema:
                return self._resolve_range_schema(schema, default_sheet)
            if self._is_cell_object(schema):
                return self._resolve_cell_object(schema, default_sheet)
            return self._resolve_dict_schema(schema, default_sheet)

        raise InvalidSchemaError(
            f"Schema value must be str, list, or dict. Got: {type(schema).__name__!r}"
        )

    def _resolve_cell_object(
        self, schema: dict, default_sheet: str | int | None
    ) -> CellValue:
        cell_ref: str = schema["cell"]
        sheet_override = schema.get("sheet")
        if sheet_override is not None:
            # sheet キーが明示されている場合はセル参照のシートプレフィックスより優先
            addr = parse_cell_ref(cell_ref)
            sheet_name = self._resolve_sheet(sheet_override)
            ws = self._get_worksheet(sheet_name)
            value = self._cell_value(ws, addr.row, addr.col, cell_ref)
            if self._empty_cell == "empty" and value is None:
                return ""
            return value
        return self._resolve_cell_ref_value(cell_ref, default_sheet)

    def _resolve_cell_ref_value(
        self, ref: str, default_sheet: str | int | None
    ) -> CellValue:
        addr = parse_cell_ref(ref)
        sheet_name = (
            self._resolve_sheet(addr.sheet)
            if addr.sheet
            else (
                self._resolve_sheet(default_sheet)
                if default_sheet is not None
                else self._default_sheet_name()
            )
        )
        ws = self._get_worksheet(sheet_name)
        value = self._cell_value(ws, addr.row, addr.col, ref)
        if self._empty_cell == "empty" and value is None:
            return ""
        return value

    def _resolve_list_schema(
        self, schema: list, default_sheet: str | int | None
    ) -> list:
        if len(schema) != 1 or not isinstance(schema[0], str):
            raise InvalidSchemaError(
                "List schema must contain exactly one range reference string. "
                f"Got: {schema!r}"
            )
        range_ref = schema[0]
        addr = parse_range_ref(range_ref)
        sheet_name = (
            self._resolve_sheet(addr.sheet)
            if addr.sheet
            else (
                self._resolve_sheet(default_sheet)
                if default_sheet is not None
                else self._default_sheet_name()
            )
        )
        ws = self._get_worksheet(sheet_name)
        rows = self._read_range_2d(ws, addr)

        flat = [cell for row in rows for cell in row]

        if self._empty_cell == "omit":
            flat = [v for v in flat if v is not None]
        elif self._empty_cell == "empty":
            flat = ["" if v is None else v for v in flat]

        return flat

    def _resolve_range_schema(
        self, schema: dict, default_sheet: str | int | None
    ) -> list:
        if "$range" not in schema:
            raise InvalidSchemaError("$schema requires $range to be specified.")
        if "$schema" not in schema:
            raise InvalidSchemaError("$range requires $schema to be specified.")

        range_ref = schema["$range"]
        col_map: dict[str, int] = schema["$schema"]
        direction: str = schema.get("$direction", "row")
        skip_empty: bool = schema.get("$skip_empty", False)

        if direction not in ("row", "column"):
            raise InvalidSchemaError(
                f"$direction must be 'row' or 'column'. Got: {direction!r}"
            )

        addr = parse_range_ref(range_ref)
        sheet_name = (
            self._resolve_sheet(addr.sheet)
            if addr.sheet
            else (
                self._resolve_sheet(default_sheet)
                if default_sheet is not None
                else self._default_sheet_name()
            )
        )
        ws = self._get_worksheet(sheet_name)
        rows_2d = self._read_range_2d(ws, addr)

        if direction == "row":
            items = []
            for row_data in rows_2d:
                if skip_empty and all(v is None for v in row_data):
                    continue
                obj = {}
                for field, idx in col_map.items():
                    val = row_data[idx] if idx < len(row_data) else None
                    if self._empty_cell == "omit" and val is None:
                        continue
                    elif self._empty_cell == "empty" and val is None:
                        val = ""
                    obj[field] = val
                items.append(obj)
            return items
        else:  # column
            num_rows = len(rows_2d)
            num_cols = len(rows_2d[0]) if rows_2d else 0
            items = []
            for col_idx in range(num_cols):
                col_data = [rows_2d[r][col_idx] for r in range(num_rows)]
                if skip_empty and all(v is None for v in col_data):
                    continue
                obj = {}
                for field, idx in col_map.items():
                    val = col_data[idx] if idx < len(col_data) else None
                    if self._empty_cell == "omit" and val is None:
                        continue
                    elif self._empty_cell == "empty" and val is None:
                        val = ""
                    obj[field] = val
                items.append(obj)
            return items

    def _resolve_dict_schema(
        self, schema: dict, default_sheet: str | int | None
    ) -> dict:
        result = {}
        for key, value in schema.items():
            resolved_value = self._resolve_schema(value, default_sheet)

            # Dynamic key: both key and value-expression look like bare cell refs
            # AND the resolved value is used, while the key's cell value becomes key name.
            if (
                is_bare_cell_ref(key)
                and isinstance(value, str)
                and is_bare_cell_ref(value)
            ):
                actual_key = self._resolve_cell_ref_value(key, default_sheet)
                actual_key = str(actual_key) if actual_key is not None else key
            else:
                actual_key = key

            if self._empty_cell == "omit" and resolved_value is None:
                continue

            result[actual_key] = resolved_value
        return result
